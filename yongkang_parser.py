"""
永康儀表板 - 資料解析模組
永康日報表是舊版 .xls 格式，用 xlrd 讀取（不是 openpyxl）。
v4 改動：YoY 去年同期改用「年度各月累計 + 當月 sheet 月累」組合。
"""

import io
import os
import re
import json
from datetime import date, datetime
from collections import defaultdict, OrderedDict

import xlrd

from drive_reader import download_file, DAILY_REPORT_FOLDER_ID

DATA_FILE = "yongkang_dashboard_data.json"
MONTHS = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']

TRACKED_MODELS = {'CR-V', 'HRV', 'FIT', 'CIVIC', 'PRELUDE'}

TEAM_STRUCTURE = {
    '永康一課': ['謝岱宏', '林昭吾', '林明昌', '陳靖玟', '林駿軒'],
    '永康二課': ['洪嘉綺', '林祐諄', '王少佟', '李霽杰', '陳冠甫'],
    '永康三課': ['蘇嵩閔', '陳建志', '黃湘婷', '江念澤', '王裕維', '張詠竣'],
}
PERSON_TO_DEPT = {p: dept for dept, members in TEAM_STRUCTURE.items() for p in members}
TEAM_ORDER = [p for members in TEAM_STRUCTURE.values() for p in members]

BLACKLIST_SUBSTR = ['合計', '月累', '課', '營業', '永康', '公司', '領牌', '訂單']

KPI_COL_REG     = 37
KPI_COL_YISHI   = 43
KPI_COL_BINGSHI = 44
KPI_COL_PJIAN   = 162

DEPT_ROW_KEYWORDS = {
    '一課月累': '永康一課',
    '二課月累': '永康二課',
    '三課月累': '永康三課',
}

RENEWAL_FILE_ID = "1-4og2dN4QI-E2XV-yD3nJ4z5v0JS7Pph"

# 永康114年度累計報表 sheet 名稱
YK_114_YEAR_SHEET = "114年度累計報表"


def is_person_name(name):
    if not name or not str(name).strip():
        return False
    name = str(name).strip()
    if any(s in name for s in BLACKLIST_SUBSTR):
        return False
    if len(name) < 2 or len(name) > 5:
        return False
    return True


def normalize_model(name):
    if name is None:
        return None
    n = str(name).strip().upper().replace(' ', '').replace('\n', '')
    mapping = {
        'CR-V': 'CR-V', 'CRV': 'CR-V', 'HR-V': 'HRV', 'HRV': 'HRV',
        'FIT': 'FIT', 'CITY': 'CITY', 'CIVIC': 'CIVIC',
        'PRELUDE': 'PRELUDE', 'ODYSSEY': 'ODYSSEY', 'ACCORD': 'ACCORD',
    }
    return mapping.get(n, n)


def _parse_filename_date(filename):
    m = re.search(r'\d{3}\s*([0-9\-\s]+)', filename)
    if not m:
        return (0, 0)
    raw = m.group(1).replace(' ', '').strip('-')
    if not raw:
        return (0, 0)
    parts = raw.split('-')
    try:
        if len(parts) == 1:
            digits = parts[0]
            if len(digits) < 3:
                return (0, 0)
            if len(digits) == 3:
                mm, dd = digits[0], digits[1:]
            else:
                mm, dd = digits[:2], digits[2:4]
            return (int(mm), int(dd))
        else:
            end = parts[-1]
            start = parts[0]
            if len(end) >= 4:
                mm, dd = end[:2], end[2:4]
            elif len(end) <= 2:
                mm = start[:2] if len(start) >= 3 else start[:1]
                dd = end
            else:
                mm, dd = end[:1], end[1:]
            return (int(mm), int(dd))
    except (ValueError, IndexError):
        return (0, 0)


def find_latest_115_file(folder_id):
    from drive_reader import get_drive_service
    service = get_drive_service()
    query = f"'{folder_id}' in parents and name contains '永康日報表115' and trashed = false"
    resp = service.files().list(q=query, pageSize=300, fields="files(id, name)").execute()
    files = resp.get("files", [])
    if not files:
        return None
    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date.sort(key=lambda x: x[1], reverse=True)
    return files_with_date[0][0]


def find_closest_114_file(folder_id, target_month, target_day):
    from drive_reader import get_drive_service
    service = get_drive_service()
    query = f"'{folder_id}' in parents and name contains '永康日報表114' and trashed = false"
    resp = service.files().list(q=query, pageSize=300, fields="files(id, name)").execute()
    files = resp.get("files", [])
    if not files:
        return None, "Drive查詢114年永康檔案回傳空清單"

    def day_ord(mm, dd):
        return mm * 31 + dd

    target_ord = day_ord(target_month, target_day)
    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date = [(f, md) for f, md in files_with_date if md[0] != 0]
    if not files_with_date:
        return None, "114年永康檔案都解析不出日期"
    files_with_date.sort(key=lambda x: abs(day_ord(*x[1]) - target_ord))
    best_file, (mm, dd) = files_with_date[0]
    return (best_file, mm, dd), None


def _resolve_sheet_name(wb, sheet_name):
    target = sheet_name.strip()
    for name in wb.sheet_names():
        if name.strip() == target:
            return name
    return None


def sheet_to_grid(wb, sheet_name):
    actual_name = _resolve_sheet_name(wb, sheet_name)
    if actual_name is None:
        return []
    sh = wb.sheet_by_name(actual_name)
    grid = []
    for r in range(sh.nrows):
        grid.append(tuple(sh.cell_value(r, c) for c in range(sh.ncols)))
    return grid


def grid_get(grid, row, col):
    r_idx, c_idx = row - 1, col - 1
    if r_idx < 0 or r_idx >= len(grid):
        return None
    row_tuple = grid[r_idx]
    if c_idx < 0 or c_idx >= len(row_tuple):
        return None
    v = row_tuple[c_idx]
    return v if v != '' else None


def find_groups(grid, header_row, sub_row, start_col, end_col_exclusive):
    starts = []
    for c in range(start_col, end_col_exclusive):
        v = grid_get(grid, header_row, c)
        if v is not None and str(v).strip():
            starts.append((c, str(v).strip()))
    groups = []
    for i, (c, name) in enumerate(starts):
        next_c = starts[i + 1][0] if i + 1 < len(starts) else end_col_exclusive
        cumcol = None
        for cc in range(c, next_c):
            sv = grid_get(grid, sub_row, cc)
            if sv is not None and '月累' in str(sv):
                cumcol = cc
                break
        if cumcol is not None:
            model = normalize_model(name)
            if model in TRACKED_MODELS:
                groups.append((model, c, cumcol))
    return groups


def parse_month_sheet(wb, sheet_name):
    grid = sheet_to_grid(wb, sheet_name)
    if not grid:
        return {}
    reg_groups = find_groups(grid, 3, 4, 2, 37)
    ord_groups = find_groups(grid, 3, 4, 54, 87)

    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        result.setdefault(name, {})
        for model, gc, cumcol in reg_groups:
            v = grid_get(grid, r, cumcol)
            v = v if isinstance(v, (int, float)) else 0
            result[name].setdefault(model, {}).setdefault('領牌', 0)
            result[name][model]['領牌'] += v
        for model, gc, cumcol in ord_groups:
            v = grid_get(grid, r, cumcol)
            v = v if isinstance(v, (int, float)) else 0
            result[name].setdefault(model, {}).setdefault('訂單', 0)
            result[name][model]['訂單'] += v
    return result


CRV_TRIM_ORDER_COLS = {'CR-V(PET)': [59, 60], 'CR-V(e:HEV)': [61, 62]}


def parse_crv_trim_split(wb, sheet_name):
    grid = sheet_to_grid(wb, sheet_name)
    if not grid:
        return {}
    sub_row = 4
    ehev_cols = CRV_TRIM_ORDER_COLS['CR-V(e:HEV)']
    label_es = grid_get(grid, sub_row, ehev_cols[0])
    label_ep = grid_get(grid, sub_row, ehev_cols[1])
    is_new_layout = (
        label_es and str(label_es).strip().upper() == 'ES'
        and label_ep and str(label_ep).strip().upper() == 'EP'
    )
    if not is_new_layout:
        return {}
    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        result.setdefault(name, {})
        for label, cols in CRV_TRIM_ORDER_COLS.items():
            total = 0
            for c in cols:
                v = grid_get(grid, r, c)
                if isinstance(v, (int, float)):
                    total += v
            result[name][label] = total
    return result


def parse_year_summary(wb, sheet_name):
    """保留供相容，讀年度累計col27。"""
    grid = sheet_to_grid(wb, sheet_name)
    if not grid:
        return {}
    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        val = grid_get(grid, r, 27)
        if isinstance(val, (int, float)):
            result[name] = result.get(name, 0) + val
    return result


def parse_year_by_month(wb, sheet_name, up_to_month):
    """讀114年度累計報表各月領牌加總（1 ~ up_to_month）。
    第N月領牌 = col(1+N*2)，確認：1月=col3, 2月=col5。
    回傳 {person: 加總領牌}
    """
    grid = sheet_to_grid(wb, sheet_name)
    if not grid:
        return {}
    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        total = 0
        for m in range(1, up_to_month + 1):
            col = 1 + m * 2
            v = grid_get(grid, r, col)
            total += int(v) if isinstance(v, (int, float)) else 0
        result[name] = total
    return result


def parse_month_reg_total(wb, sheet_name):
    """讀某月 sheet 的月累計領牌，回傳 {person: 領牌月累}。備援用。"""
    r = parse_month_sheet(wb, sheet_name)
    result = {}
    for person, models in r.items():
        total = sum(v.get('領牌', 0) for v in models.values())
        if total > 0:
            result[person] = total
    return result


def _safe_pct(n, d):
    return round(n / d, 4) if d and d > 0 else None


def _build_kpi_row(reg, yi, bing, pjia):
    full = yi + bing
    base = reg
    return {
        'reg': reg, 'base': base, 'full': full, 'yi': yi,
        'acc_t': int(pjia), 'acc_per': round(pjia / reg) if reg > 0 else 0,
        '全險比': _safe_pct(full, base), '乙式比': _safe_pct(yi, base),
    }


def parse_month_kpi(wb, sheet_name):
    grid = sheet_to_grid(wb, sheet_name)
    if not grid:
        return {'person': {}, 'dept': {}}

    person_data = {}
    dept_data   = {}

    for r in range(1, len(grid) + 1):
        raw = grid_get(grid, r, 1)
        if not raw or not str(raw).strip():
            continue
        name = str(raw).strip()

        def gv(col):
            v = grid_get(grid, r, col) or 0
            return int(round(v)) if isinstance(v, (int, float)) else 0

        reg  = gv(KPI_COL_REG)
        yi   = gv(KPI_COL_YISHI)
        bing = gv(KPI_COL_BINGSHI)
        pjia = gv(KPI_COL_PJIAN)

        matched_dept = DEPT_ROW_KEYWORDS.get(name)
        if matched_dept:
            dept_data[matched_dept] = _build_kpi_row(reg, yi, bing, pjia)
            continue

        if not is_person_name(name):
            continue

        person_data[name] = _build_kpi_row(reg, yi, bing, pjia)

    return {'person': person_data, 'dept': dept_data}


def read_renewal_progress():
    import openpyxl
    from datetime import date as _date

    content = download_file(RENEWAL_FILE_ID)
    wb_renew = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    curr_month = _date.today().month
    target_name = f"115.{curr_month:02d}續保"
    sh = None
    for sname in wb_renew.sheetnames:
        if sname.strip() == target_name:
            sh = wb_renew[sname]
            break
    if sh is None:
        for sname in reversed(wb_renew.sheetnames):
            if '續保' in sname:
                sh = wb_renew[sname]
                break
    if sh is None:
        sh = wb_renew.worksheets[0]

    DEPT_EXACT = {'一課': '永康一課', '二課': '永康二課', '三課': '永康三課'}
    person_data = {}
    dept_data   = {}
    total_data  = {}

    for row in sh.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        raw_name = str(row[0]).strip()
        if not raw_name:
            continue
        try:
            boushu = int(round(float(row[1]))) if row[1] not in (None, '') else 0
            yugu   = int(round(float(row[2]))) if row[2] not in (None, '') else 0
            yishou = int(round(float(row[3]))) if row[3] not in (None, '') else 0
        except (ValueError, TypeError):
            continue
        rate = round(yishou / boushu * 100, 1) if boushu > 0 else 0.0
        row_data = {'母數': boushu, '預估': yugu, '已收': yishou, '續保率': rate}

        if raw_name in DEPT_EXACT:
            dept_data[DEPT_EXACT[raw_name]] = row_data
        elif '合計' in raw_name:
            total_data = row_data
        elif is_person_name(raw_name):
            person_data[raw_name] = row_data

    wb_renew.close()
    return {'person': person_data, 'dept': dept_data, 'total': total_data}


def compute_dept_totals_scalar(person_value_dict):
    totals = {dept: 0 for dept in TEAM_STRUCTURE}
    for p, v in person_value_dict.items():
        dept = PERSON_TO_DEPT.get(p)
        if dept:
            totals[dept] += v
    return totals


def compute_dept_totals_by_model(person_model_dict):
    totals = {dept: defaultdict(int) for dept in TEAM_STRUCTURE}
    for p, models in person_model_dict.items():
        dept = PERSON_TO_DEPT.get(p)
        if dept:
            for model, v in models.items():
                totals[dept][model] += v
    return {dept: dict(models) for dept, models in totals.items()}


def sort_by_total_desc(d):
    return dict(sorted(d.items(), key=lambda item: sum(item[1].values()) if isinstance(item[1], dict) else item[1], reverse=True))


HISTORY_DRIVE_FILENAME = "yongkang_order_history.json"
SEED_HISTORY_FILE = "yongkang_order_history_seed.json"


def load_order_history():
    from drive_json_store import load_json_from_drive
    history = {}
    if os.path.exists(SEED_HISTORY_FILE):
        with open(SEED_HISTORY_FILE, "r", encoding="utf-8") as f:
            history.update(json.load(f))
    history.update(load_json_from_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME))
    for snap in history.values():
        for models in snap.values():
            models.pop('CR-V', None)
    return history


def save_order_history(history):
    from drive_json_store import save_json_to_drive
    save_json_to_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME, history)


MONTH_LAST_DAY_2026 = {
    '1月': (1, 31), '2月': (2, 28), '3月': (3, 31), '4月': (4, 30),
    '5月': (5, 31), '6月': (6, 30), '7月': (7, 31), '8月': (8, 31),
    '9月': (9, 30), '10月': (10, 31), '11月': (11, 30), '12月': (12, 31),
}

DAILY_VALUE_MODELS = {'CR-V(PET)', 'CR-V(e:HEV)'}


def compute_last_order_tracking(history, today_str):
    dates_sorted = sorted(history.keys())
    people, models = set(), set()
    for snap in history.values():
        for p, md in snap.items():
            people.add(p)
            models.update(md.keys())

    by_month = OrderedDict()
    for d in dates_sorted:
        by_month.setdefault(d[:7], []).append(d)

    global_first_date = dates_sorted[0] if dates_sorted else None
    only_one_snapshot_ever = len(dates_sorted) <= 1

    today = date.fromisoformat(today_str)
    result = {}
    for p in people:
        result[p] = {}
        for model in models:
            last_date = None
            if model in DAILY_VALUE_MODELS:
                for d in dates_sorted:
                    cur_val = history[d].get(p, {}).get(model, 0)
                    if cur_val > 0:
                        last_date = d
            else:
                if only_one_snapshot_ever:
                    continue
                for ym, ds in by_month.items():
                    prev_val = None
                    for d in ds:
                        cur_val = history[d].get(p, {}).get(model, 0)
                        if prev_val is None:
                            if cur_val > 0 and d != global_first_date:
                                last_date = d
                        else:
                            if cur_val > prev_val:
                                last_date = d
                        prev_val = cur_val
            if last_date is None:
                continue
            days_since = (today - date.fromisoformat(last_date)).days
            result[p][model] = {"last_order_date": last_date, "days_since": days_since, "approx": False}
    return result


def fill_fallback_from_monthly(last_order_tracking, month_sheets_cache, months_to_sum, today_str):
    today = date.fromisoformat(today_str)
    people_models_with_data = set()
    for m in months_to_sum:
        for p, models in month_sheets_cache.get(m, {}).items():
            for model, vals in models.items():
                if vals.get('訂單', 0) > 0:
                    people_models_with_data.add((p, model))

    for p, model in people_models_with_data:
        if model in DAILY_VALUE_MODELS:
            continue
        if last_order_tracking.get(p, {}).get(model) is not None:
            continue
        for m in reversed(months_to_sum):
            v = month_sheets_cache.get(m, {}).get(p, {}).get(model, {}).get('訂單', 0)
            if v > 0:
                mm, dd = MONTH_LAST_DAY_2026[m]
                candidate = date(today.year, mm, dd)
                if candidate > today:
                    candidate = today
                last_order_tracking.setdefault(p, {})
                last_order_tracking[p][model] = {
                    "last_order_date": candidate.isoformat(),
                    "days_since": (today - candidate).days,
                    "approx": True,
                }
                break
    return last_order_tracking


def build_daily_snapshot(wb, month_key):
    r = parse_month_sheet(wb, month_key)
    crv_split = parse_crv_trim_split(wb, month_key)
    snapshot = {}
    for person, models in r.items():
        snapshot[person] = {k: v.get('訂單', 0) for k, v in models.items()
                             if '訂單' in v and k != 'CR-V'}
    for person, split_vals in crv_split.items():
        snapshot.setdefault(person, {})
        for label, val in split_vals.items():
            snapshot[person][label] = val
    return snapshot


def reset_order_history():
    from drive_json_store import save_json_to_drive
    save_json_to_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME, {})
    return {"status": "ok", "message": "Drive上的歷史紀錄已清空，種子檔不受影響"}


def backfill_full_history(max_seconds=90, max_files=8):
    import time
    import gc
    start_time = time.time()

    from drive_reader import get_drive_service, download_file
    service = get_drive_service()
    query = f"'{DAILY_REPORT_FOLDER_ID}' in parents and name contains '永康日報表115' and trashed = false"
    resp = service.files().list(q=query, pageSize=300, fields="files(id, name)").execute()
    files = resp.get("files", [])

    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date = [(f, md) for f, md in files_with_date if md[0] != 0]
    files_with_date.sort(key=lambda x: x[1])

    history = load_order_history()
    processed = 0
    errors = []
    timed_out = False

    for f, (mm, dd) in files_with_date:
        date_str = f"2026-{mm:02d}-{dd:02d}"
        if date_str in history:
            continue
        if processed >= max_files or time.time() - start_time > max_seconds:
            timed_out = True
            break
        try:
            content = download_file(f["id"])
            wb = xlrd.open_workbook(file_contents=content)
            month_key = f"{mm}月"
            if _resolve_sheet_name(wb, month_key) is not None:
                history[date_str] = build_daily_snapshot(wb, month_key)
                processed += 1
            del content, wb
            gc.collect()
        except Exception as e:
            errors.append(f"{f['name']}: {str(e)[:100]}")

    save_order_history(history)

    return {
        "processed_this_run": processed,
        "total_files_found": len(files_with_date),
        "dates_now_in_history": len(history),
        "timed_out": timed_out,
        "done": not timed_out,
        "errors": errors[:10],
    }


def build_yongkang_data():
    file_info = find_latest_115_file(DAILY_REPORT_FOLDER_ID)
    if not file_info:
        raise RuntimeError("Drive 資料夾裡找不到115年永康日報表檔案")

    content = download_file(file_info["id"])
    wb = xlrd.open_workbook(file_contents=content)

    today = date.today()
    current_month_key = f"{today.month}月"

    ytd = defaultdict(lambda: defaultdict(lambda: {'領牌': 0, '訂單': 0}))
    months_to_sum = [m for m in MONTHS if _resolve_sheet_name(wb, m) is not None]
    month_sheets_cache = {}
    month_sheets_cache_for_tracking = {}
    for m in months_to_sum:
        r = parse_month_sheet(wb, m)
        month_sheets_cache[m] = r
        for person, models in r.items():
            for model, vals in models.items():
                ytd[person][model]['領牌'] += vals.get('領牌', 0)
                ytd[person][model]['訂單'] += vals.get('訂單', 0)

        crv_split = parse_crv_trim_split(wb, m)
        r_tracking = {}
        for person, models in r.items():
            r_tracking[person] = {k: v for k, v in models.items() if k != 'CR-V'}
        for person, split_vals in crv_split.items():
            r_tracking.setdefault(person, {})
            for label, val in split_vals.items():
                r_tracking[person][label] = {'訂單': val}
        month_sheets_cache_for_tracking[m] = r_tracking

    item1 = {p: sum(v['領牌'] for v in models.values()) for p, models in ytd.items()}
    team_total_ytd_registration = sum(item1.values())
    item1_dept_totals = compute_dept_totals_scalar(item1)

    item4 = {p: {mo: v['領牌'] for mo, v in models.items() if v['領牌'] > 0}
             for p, models in ytd.items()}
    item4 = sort_by_total_desc(item4)
    item4_dept_totals = compute_dept_totals_by_model(item4)

    cur = month_sheets_cache.get(current_month_key, {})
    cur_tracking = month_sheets_cache_for_tracking.get(current_month_key, {})
    month_progress = {}
    for p, models in cur.items():
        month_progress[p] = {
            '訂單': sum(v.get('訂單', 0) for v in models.values()),
            '領牌': sum(v.get('領牌', 0) for v in models.values()),
        }
    month_progress_dept_totals = compute_dept_totals_by_model(month_progress)

    cur_kpi = parse_month_kpi(wb, current_month_key)

    try:
        renewal = read_renewal_progress()
    except Exception as e:
        import traceback
        renewal = {"error": str(e), "trace": traceback.format_exc()[-300:]}

    today_str = today.isoformat()
    history = load_order_history()
    if cur_tracking:
        snapshot = {p: {m: v.get('訂單', 0) for m, v in models.items() if '訂單' in v}
                    for p, models in cur_tracking.items()}
        history[today_str] = snapshot
        save_order_history(history)
    last_order_tracking = compute_last_order_tracking(history, today_str)
    last_order_tracking = fill_fallback_from_monthly(
        last_order_tracking, month_sheets_cache_for_tracking, months_to_sum, today_str
    )

    item1 = {p: v for p, v in item1.items() if p in TEAM_ORDER}
    item4 = {p: v for p, v in item4.items() if p in TEAM_ORDER}
    month_progress = {p: v for p, v in month_progress.items() if p in TEAM_ORDER}
    last_order_tracking = {p: v for p, v in last_order_tracking.items() if p in TEAM_ORDER}

    # ── YoY：去年同期 = 年度各月累計（上月底）+ 當月 sheet 月累（即時）──
    yoy_comparison = None
    try:
        found, err = find_closest_114_file(DAILY_REPORT_FOLDER_ID, today.month, today.day)
        if found is None:
            yoy_comparison = {"error": err}
        else:
            last_year_file, ly_month, ly_day = found
            ly_content = download_file(last_year_file["id"])
            ly_wb = xlrd.open_workbook(file_contents=ly_content)

            # 1. 年度sheet讀上月底（1 ~ curr_month-1）各月領牌加總
            prev_month = today.month - 1
            if prev_month >= 1:
                ly_prev = parse_year_by_month(ly_wb, YK_114_YEAR_SHEET, prev_month)
            else:
                ly_prev = {}

            # 2. 年度sheet讀當月數字
            curr_col = 1 + today.month * 2
            ly_ytd_grid = sheet_to_grid(ly_wb, YK_114_YEAR_SHEET)
            curr_from_ytd = {}
            for r in range(1, len(ly_ytd_grid) + 1):
                name = grid_get(ly_ytd_grid, r, 1)
                if not is_person_name(name):
                    continue
                name = str(name).strip()
                v = grid_get(ly_ytd_grid, r, curr_col)
                curr_from_ytd[name] = int(v) if isinstance(v, (int, float)) else 0

            # 3. 當月 sheet 月累（備援）
            ly_curr_sheet = f"{today.month}月"
            ly_curr_reg = {}
            if _resolve_sheet_name(ly_wb, ly_curr_sheet) is not None:
                ly_curr_reg = parse_month_reg_total(ly_wb, ly_curr_sheet)

            del ly_content

            # 4. 合併：上月累計 + 當月（年度sheet有值優先，否則用當月sheet）
            all_people = set(ly_prev.keys()) | set(curr_from_ytd.keys()) | set(ly_curr_reg.keys())
            ly_combined = {}
            for p in all_people:
                prev = ly_prev.get(p, 0)
                curr_ytd = curr_from_ytd.get(p, 0)
                curr_sheet = ly_curr_reg.get(p, 0)
                curr_final = curr_ytd if curr_ytd > 0 else curr_sheet
                ly_combined[p] = prev + curr_final

            yoy_comparison = {
                "last_year_file": last_year_file["name"],
                "last_year_date": f"2025-{ly_month:02d}-{ly_day:02d}",
                "last_year_ytd": {k: int(v) for k, v in ly_combined.items()},
                "yoy_method": f"年度各月累計(1~{prev_month}月) + 當月sheet即時領牌",
            }
    except Exception as e:
        import traceback
        yoy_comparison = {"error": str(e), "trace": traceback.format_exc()[-500:]}

    yoy_last_year_dept_totals = None
    yoy_this_year_dept_totals = None
    if yoy_comparison and yoy_comparison.get("last_year_ytd") is not None:
        yoy_last_year_dept_totals = compute_dept_totals_scalar(yoy_comparison["last_year_ytd"])
        yoy_this_year_dept_totals = item1_dept_totals

    def to_int(x):
        try:
            return int(round(x))
        except (TypeError, ValueError):
            return x

    def clean_scalar_dict(d):
        return {k: to_int(v) for k, v in d.items()}

    def clean_model_dict(d):
        return {k: {m: to_int(v) for m, v in models.items()} for k, models in d.items()}

    item1 = clean_scalar_dict(item1)
    item1_dept_totals = clean_scalar_dict(item1_dept_totals)
    team_total_ytd_registration = to_int(team_total_ytd_registration)
    item4 = clean_model_dict(item4)
    item4_dept_totals = clean_model_dict(item4_dept_totals)
    month_progress = clean_model_dict(month_progress)
    month_progress_dept_totals = clean_model_dict(month_progress_dept_totals)
    if yoy_comparison and yoy_comparison.get("last_year_ytd") is not None:
        yoy_last_year_dept_totals = clean_scalar_dict(yoy_last_year_dept_totals)
        yoy_this_year_dept_totals = item1_dept_totals

    data = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": file_info["name"],
        "team_structure": {k: TEAM_STRUCTURE[k] for k in ['永康一課', '永康二課', '永康三課']},
        "item1_ytd_registration": item1,
        "item1_dept_totals": item1_dept_totals,
        "team_total_ytd_registration": team_total_ytd_registration,
        "item4_ytd_by_model": item4,
        "item4_dept_totals": item4_dept_totals,
        "month_progress": month_progress,
        "month_progress_dept_totals": month_progress_dept_totals,
        "last_order_tracking": last_order_tracking,
        "yoy_comparison": yoy_comparison,
        "yoy_last_year_dept_totals": yoy_last_year_dept_totals,
        "yoy_this_year_dept_totals": yoy_this_year_dept_totals,
        "month_kpi": cur_kpi,
        "renewal": renewal,
    }

    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return data


if __name__ == "__main__":
    result = build_yongkang_data()
    print(json.dumps(result, ensure_ascii=False, indent=2))
