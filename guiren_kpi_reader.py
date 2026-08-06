"""
歸仁績效指標資料讀取模組
資料來源：
1. 日報表（Drive）- 領牌台數（年度/當月）+ 當月全險/乙式/分期/配件
2. 週邊指標（Drive）- 歷史月份全險/乙式/分期/配件
3. 續保進度表（Drive）- 當月續保進度 + 年度成績（已完成月份）
4. 有望客名單（Drive）- 來店/邀約（上個月更新）
"""

import io, json, base64, logging
from datetime import datetime
from collections import defaultdict
import pytz

TZ = pytz.timezone('Asia/Taipei')

SA_DISPLAY = ['林定緯', '林適緯', '陳建道', '陳星佑', '張姉瑀', '歐陽文智', '蔡明憬']
SA_ALL     = ['林定緯', '林適緯', '劉珈微', '陳建道', '陳星佑', '張姉瑀', '歐陽文智', '蔡明憬']
C1_ALL     = ['林定緯', '林適緯', '劉珈微', '陳建道']
C2_ALL     = ['陳星佑', '蔡明憬', '張姉瑀', '歐陽文智']
COMPANY    = '巧克力汽車商業股份有限公司'

KPI_FILE_ID    = '1-8JmC6MlF-WPMsgeWqh43f_q5BiA3Q_G'  # 週邊指標
RENEW_FILE_ID  = '1-6Wmly1lKSLVOEUspwcV9TPsghdjyMC_'  # 續保進度表
DAILY_FOLDER   = '1SDP7OJ79g6WqaoDqAQEeHZwj09MHTWyf'  # 日報表資料夾

SA_MAP_KPI = {
    '定緯':'林定緯','適緯':'林適緯','珈微':'劉珈微','建道':'陳建道',
    '星佑':'陳星佑','姉瑀':'張姉瑀','文智':'歐陽文智','明憬':'蔡明憬'
}

MONTH_NAMES = ['一月','二月','三月','四月','五月','六月',
               '七月','八月','九月','十月','十一月','十二月']


def _dl_xlsx_bytes(service, file_id):
    """Drive 下載 xlsx → raw bytes（不解析，節省記憶體）"""
    req = service.files().get_media(fileId=file_id)
    return req.execute()

def _dl_xlsx(service, file_id):
    """Drive 下載 xlsx → openpyxl Workbook（read_only 省記憶體）"""
    import openpyxl
    content = _dl_xlsx_bytes(service, file_id)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    del content  # 立刻釋放原始 bytes
    return wb


def _latest_daily_file(service):
    """取得最新一份歸仁日報表的 file_id"""
    result = service.files().list(
        q=f"'{DAILY_FOLDER}' in parents and name contains '歸仁日報表' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        orderBy='modifiedTime desc',
        pageSize=1,
        fields='files(id,name)'
    ).execute()
    items = result.get('files', [])
    if not items:
        return None, None
    return items[0]['id'], items[0]['name']


def _latest_prospect_file(service):
    """取得最新一份有望客名單的 file_id"""
    result = service.files().list(
        q="name contains '歸仁有望客名單' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        orderBy='modifiedTime desc',
        pageSize=1,
        fields='files(id,name)'
    ).execute()
    items = result.get('files', [])
    if not items:
        return None, None
    return items[0]['id'], items[0]['name']


def _safe_pct(n, d):
    return round(n / d, 4) if d and d > 0 else None


def _sum_kpi(data, members):
    reg=base=full=yi=loan=acc_t=prem_t=0
    for sa in members:
        d = data.get(sa, {})
        reg   += d.get('reg', 0)
        base  += d.get('base', 0)
        full  += d.get('full', 0)
        yi    += d.get('yi', 0)
        loan  += d.get('loan', 0)
        acc_t += d.get('acc_t', 0)
        prem_t+= d.get('prem_t', 0)
    return {
        'reg': reg, 'base': base,
        'full': full, 'yi': yi, 'loan': loan,
        'acc_t': acc_t, 'prem_t': prem_t,
        'acc_per': round(acc_t/reg) if reg > 0 else 0,
        '全險比': _safe_pct(full, base),
        '乙式比': _safe_pct(yi, base),
        '分期比': _safe_pct(loan, base),
    }


# ─────────────────────────────────────────
# 1. 日報表：領牌台數（年度累計 + 當月）
# ─────────────────────────────────────────
def read_daily(service):
    """
    回傳：
    {
      'ytd':  { sa: {'ord': N, 'reg': N}, '一課': {...}, '二課': {...}, '歸仁': {...} },
      'curr_month': { sa: {'ord': N, 'reg': N}, ... },
      'curr_month_num': 8,   # 當月月份
      'last_updated': '08/04',
    }
    """
    now = datetime.now(TZ)
    curr_m = now.month

    file_id, title = _latest_daily_file(service)
    if not file_id:
        return {}

    wb = _dl_xlsx(service, file_id)
    ws = wb['115年度']

    # col mapping: B=1月ord, C=1月reg, D=2月ord, E=2月reg...
    # 0-indexed: 1月ord=col1, 1月reg=col2, 2月ord=col3, 2月reg=col4...
    # month m: ord_col = 1+(m-1)*2, reg_col = 2+(m-1)*2
    # 年度累計: ord = col25, reg = col26

    SA_NAMES_DAILY = SA_ALL + ['一課合計','二課合計']

    ytd = {}
    curr = {}

    for row in ws.iter_rows(values_only=True):
        name = str(row[0] or '').strip()
        if name not in SA_NAMES_DAILY + ['一課合計','二課合計','總月累']:
            continue

        ytd_ord = row[25] or 0
        ytd_reg = row[26] or 0

        # 當月
        curr_ord_col = 1 + (curr_m - 1) * 2
        curr_reg_col = 2 + (curr_m - 1) * 2
        cm_ord = row[curr_ord_col] or 0
        cm_reg = row[curr_reg_col] or 0

        if name in SA_ALL:
            ytd[name]  = {'ord': ytd_ord, 'reg': ytd_reg}
            curr[name] = {'ord': cm_ord,  'reg': cm_reg}
        elif name == '一課合計':
            ytd['歸仁一課']  = {'ord': ytd_ord, 'reg': ytd_reg}
            curr['歸仁一課'] = {'ord': cm_ord,  'reg': cm_reg}
        elif name == '二課合計':
            ytd['歸仁二課']  = {'ord': ytd_ord, 'reg': ytd_reg}
            curr['歸仁二課'] = {'ord': cm_ord,  'reg': cm_reg}
        elif name == '總月累':
            ytd['歸仁據點']  = {'ord': ytd_ord, 'reg': ytd_reg}
            curr['歸仁據點'] = {'ord': cm_ord,  'reg': cm_reg}

    return {
        'ytd': ytd,
        'curr': curr,
        'curr_month': curr_m,
        'last_updated': title,
    }


# ─────────────────────────────────────────
# 2. 週邊指標：全險比/乙式比/分期比/配件
#    歷史月份 + 當月從日報表抓
# ─────────────────────────────────────────
def read_kpi(service, curr_month):
    """
    回傳 per-SA per-period:
    {
      sa: {
        'ytd':  { reg, base, full, yi, loan, acc_t, acc_per, 全險比, 乙式比, 分期比 },
        'curr': { ... }   # 當月從月份sheet抓
      }
    }
    """
    now = datetime.now(TZ)
    wb = _dl_xlsx(service, KPI_FILE_ID)

    # 歷史完整月份：週邊指標月份 sheet
    # 當月：週邊指標當月 sheet（可能未更新，但先抓）
    sa_data = defaultdict(lambda: defaultdict(lambda:{
        'reg':0,'acc':0,'yi':0,'bing':0,'full':0,'loan':0,
        'rental':0,'no_ins':0,'base':0,'prem':0
    }))

    # 只讀 Q1/Q2/當月 sheet，降低記憶體
    target_sheets = ['2026.Q1', '2026.Q2', f'2026.{curr_month:02d}\u6708']
    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue
        if sname == '2026.Q1':
            period = 'Q1'
        elif sname == '2026.Q2':
            period = 'Q2'
        else:
            try:
                part = sname.replace('2026.','').replace('\u6708','')
                m = int(part)
            except:
                continue
            period = f'M{m}'
        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            sa_abbr  = str(row[1] or '').strip()
            customer = str(row[2] or '').strip()
            acc      = row[3]
            loan_co  = str(row[4] or '').strip()
            ins_type = str(row[6] or '').strip()
            premium  = row[7]
            if sa_abbr not in SA_MAP_KPI: continue
            sa = SA_MAP_KPI[sa_abbr]
            if COMPANY in customer: continue
            is_rental = (customer == '租車')
            sa_data[sa][period]['reg'] += 1
            if isinstance(acc, (int,float)): sa_data[sa][period]['acc'] += acc
            if isinstance(premium,(int,float)): sa_data[sa][period]['prem'] += premium
            if is_rental: sa_data[sa][period]['rental'] += 1; continue
            sa_data[sa][period]['base'] += 1
            if loan_co == '元大': sa_data[sa][period]['loan'] += 1
            if ins_type == '乙式': sa_data[sa][period]['yi']+=1; sa_data[sa][period]['full']+=1
            elif ins_type == '丙式': sa_data[sa][period]['bing']+=1; sa_data[sa][period]['full']+=1
            elif ins_type == '甲式': sa_data[sa][period]['full']+=1
            else: sa_data[sa][period]['no_ins']+=1

    # 整合：YTD = 已完整月份（不含當月）
    completed_months = list(range(1, curr_month))  # 已完整月份
    all_months = list(range(1, curr_month + 1))    # 含當月

    result = {}
    for sa in SA_ALL:
        ytd = {k:0 for k in ['reg','base','full','yi','loan','acc_t','prem_t']}
        curr = {k:0 for k in ['reg','base','full','yi','loan','acc_t','prem_t']}
        for qp in ['Q1','Q2']:
            d = sa_data[sa][qp]
            ytd['reg']   += d['reg'];  ytd['base']  += d['base']
            ytd['full']  += d['full']; ytd['yi']    += d['yi']
            ytd['loan']  += d['loan']; ytd['acc_t'] += d['acc']
            ytd['prem_t']+= d['prem']
        if curr_month > 6:
            dm = sa_data[sa][f'M{curr_month}']
            for k in ['reg','base','full','yi','loan']:
                ytd[k] += dm[k]
            ytd['acc_t']  += dm['acc']
            ytd['prem_t'] += dm['prem']
        curr.update({'reg':dm['reg'],'base':dm['base'],'full':dm['full'],
                     'yi':dm['yi'],'loan':dm['loan'],'acc_t':dm['acc'],'prem_t':dm['prem']})
        result[sa] = {
            'ytd': {
                **ytd, 'acc_per': round(ytd['acc_t']/ytd['reg']) if ytd['reg']>0 else 0,
                '全險比': _safe_pct(ytd['full'],ytd['base']),
                '乙式比': _safe_pct(ytd['yi'],ytd['base']),
                '分期比': _safe_pct(ytd['loan'],ytd['base']),
            },
            'curr': {
                **curr, 'acc_per': round(curr['acc_t']/curr['reg']) if curr['reg']>0 else 0,
                '全險比': _safe_pct(curr['full'],curr['base']),
                '乙式比': _safe_pct(curr['yi'],curr['base']),
                '分期比': _safe_pct(curr['loan'],curr['base']),
            }
        }
    wb.close()
    gc.collect()
    return result


# ─────────────────────────────────────────
# 3. 續保進度表
# ─────────────────────────────────────────
def read_renew(service, curr_month):
    import gc
    """
    回傳：
    {
      'curr_month': {  # 當月進度
        sa: { '整體': {den,num,rate,target_rate}, '首年': {...}, '首年車體': {...} }
      },
      'ytd': {  # 已完成月份加總年度成績
        sa: { '整體': {den,num,rate}, '首年車體': {...}, '首年': {...} }
      }
    }
    欄位對應（2026續保成績 sheet）：
      col0=營業員, col1=YTD母數, col2=YTD實績, col3=比例
      逐月: col8=1月母, col9=1月實, col10=2月母, col11=2月實 ...
    """
    import openpyxl
    wb = _dl_xlsx(service, RENEW_FILE_ID)

    # ── 當月進度 ──
    month_sheet = f'115.{str(curr_month).zfill(2)}月續保'
    curr_data = {}
    if month_sheet in wb.sheetnames:
        ws = wb[month_sheet]
        rows = list(ws.iter_rows(values_only=True))
        # 欄位: 0=營業員,1=母數,2=預估,3=已收,4=預估率,5=續保率,6=進度表
        #       7=首年母,8=首年預估,9=首年已收,10=首年預估率,11=首年率,12=首年進度
        #       13=首年車體母,14=首年車體預估,15=首年車體已收,16=首車體預估率,17=首車體率
        for row in rows:
            sa = str(row[0] or '').strip()
            if sa not in SA_ALL + ['歸仁一課','歸仁二課','合計','劉宗鑫']: continue
            display_sa = sa if sa in SA_ALL else sa.replace('合計','歸仁據點')
            curr_data[display_sa] = {
                '整體':    {'den': row[1] or 0, 'num': row[3] or 0,
                            'rate': row[5], 'pred_rate': row[4], 'progress': row[6]},
                '首年':    {'den': row[7] or 0,  'num': row[9] or 0,
                            'rate': row[11], 'pred_rate': row[10]},
                '首年車體':{'den': row[13] or 0, 'num': row[15] or 0,
                            'rate': row[17], 'pred_rate': row[16]},
            }

    # ── 年度成績（已完成月份加總）──
    completed = list(range(1, curr_month))  # 不含當月

    ytd_data = {}
    for section_name, sheet_key in [
        ('整體',    '2026續保成績'),
        ('首年車體', '2026首年車體續保'),
        ('首年',    '2026首年續保'),
    ]:
        if sheet_key not in wb.sheetnames: continue
        ws = wb[sheet_key]
        for row in ws.iter_rows(values_only=True):
            sa = str(row[0] or '').strip()
            if sa not in SA_ALL + ['歸仁一課','歸仁二課','歸仁據點','劉宗鑫']: continue

            # 逐月欄：col8=1月母, col9=1月實, col10=2月母, col11=2月實...
            ytd_num = ytd_den = 0
            for m in completed:
                den_col = 8 + (m-1)*2
                num_col = 9 + (m-1)*2
                if den_col < len(row):
                    ytd_den += (row[den_col] or 0)
                    ytd_num += (row[num_col] or 0)

            if sa not in ytd_data: ytd_data[sa] = {}
            ytd_data[sa][section_name] = {
                'den': ytd_den,
                'num': ytd_num,
                'rate': _safe_pct(ytd_num, ytd_den),
                'completed_months': completed,
            }

    if 'wb' in dir():
        try: wb.close()
        except: pass
    gc.collect()
    return {'curr': curr_data, 'ytd': ytd_data}


# ─────────────────────────────────────────
# 4. 有望客名單
# ─────────────────────────────────────────
def read_prospect(service):
    """
    回傳：{ sa: {'walk':N,'walkC':N,'inv':N,'invC':N} }
    （上月為止累計，因為名單每月更新一次）
    """
    from collections import defaultdict
    file_id, _ = _latest_prospect_file(service)
    if not file_id: return {}

    wb = _dl_xlsx(service, file_id)
    ws = wb.active

    SUCCESS = {'訂單','領牌','交車','退購'}
    EXCL_MEDIA = {'來廠','外展'}

    result = defaultdict(lambda:{'walk':0,'walkC':0,'inv':0,'invC':0})
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1: continue
        sa = row[5]; status = row[6]
        source = row[14]; media = row[15]
        if not sa or sa not in SA_ALL: continue
        is_ok = status in SUCCESS
        if source == '自然來店':
            result[sa]['walk'] += 1
            if is_ok: result[sa]['walkC'] += 1
        elif source == '邀約來店' and str(media or '') not in EXCL_MEDIA:
            result[sa]['inv'] += 1
            if is_ok: result[sa]['invC'] += 1
    if 'wb' in dir():
        try: wb.close()
        except: pass
    gc.collect()
    return dict(result)


# ─────────────────────────────────────────
# 主入口：整合所有資料
# ─────────────────────────────────────────
def get_guiren_kpi(service):
    """
    整合所有資料，回傳儀表板用 JSON：
    {
      'meta': { 'curr_month': 8, 'updated_at': '...' },
      'entities': {
        '歸仁據點': { 'ytd': {...}, 'curr': {...}, 'renew': {...} },
        '歸仁一課': { ... },
        '歸仁二課': { ... },
        '林定緯':   { ... },
        ...
      }
    }
    """
    now = datetime.now(TZ)
    curr_month = now.month

    # 讀取各資料源
    daily   = read_daily(service)
    kpi     = read_kpi(service, curr_month)
    renew   = read_renew(service, curr_month)
    prospect= read_prospect(service)

    ytd_map  = daily.get('ytd',  {})
    curr_map = daily.get('curr', {})

    # 聚合函式
    def build_entity(members, entity_key, is_depot=False):
        # 領牌
        ytd_reg  = sum(ytd_map.get(sa,{}).get('reg',0) for sa in members)
        curr_reg = sum(curr_map.get(sa,{}).get('reg',0) for sa in members)
        ytd_ord  = sum(ytd_map.get(sa,{}).get('ord',0) for sa in members)
        curr_ord = sum(curr_map.get(sa,{}).get('ord',0) for sa in members)

        # 週邊指標聚合
        def agg_kpi(period_key):
            reg=base=full=yi=loan=acc_t=prem_t=0
            for sa in members:
                d = kpi.get(sa, {}).get(period_key, {})
                reg   += d.get('reg',0);  base  += d.get('base',0)
                full  += d.get('full',0); yi    += d.get('yi',0)
                loan  += d.get('loan',0); acc_t += d.get('acc_t',0)
                prem_t+= d.get('prem_t',0)
            return {
                'reg': reg, 'base': base, 'full': full, 'yi': yi,
                'loan': loan, 'acc_t': acc_t, 'prem_t': prem_t,
                'acc_per': round(acc_t/reg) if reg>0 else 0,
                '全險比': _safe_pct(full,base),
                '乙式比': _safe_pct(yi,base),
                '分期比': _safe_pct(loan,base),
            }

        # 有望客聚合
        def agg_pros(members_list, include_hsin=False):
            walk=walkC=inv=invC=0
            for sa in members_list:
                p = prospect.get(sa,{})
                walk+=p.get('walk',0); walkC+=p.get('walkC',0)
                inv+=p.get('inv',0);   invC+=p.get('invC',0)
            if include_hsin:  # 據點加劉宗鑫的邀約1件
                inv += 1
            return {'walk':walk,'walkC':walkC,'inv':inv,'invC':invC,
                    'walk_rate': _safe_pct(walkC,walk),
                    'inv_rate':  _safe_pct(invC,inv)}

        # 續保
        renew_key = entity_key if entity_key in renew['curr'] else None
        if entity_key == '歸仁據點':
            renew_key = '合計'

        kpi_ytd  = agg_kpi('ytd')
        kpi_curr = agg_kpi('curr')
        pros     = agg_pros(members, include_hsin=is_depot)

        entity = {
            'ytd': {
                'reg': ytd_reg, 'ord': ytd_ord,
                **{k: kpi_ytd[k] for k in ['base','full','yi','loan','acc_t','acc_per','全險比','乙式比','分期比']},
                **pros,
            },
            'curr': {
                'reg': curr_reg, 'ord': curr_ord,
                **{k: kpi_curr[k] for k in ['base','full','yi','loan','acc_t','acc_per','全險比','乙式比','分期比']},
            },
            'renew': {
                'curr':  renew['curr'].get(renew_key or entity_key, {}),
                'ytd':   renew['ytd'].get(renew_key or entity_key, {}),
            }
        }
        return entity

    entities = {}
    entities['歸仁據點'] = build_entity(C1_ALL+C2_ALL, '歸仁據點', is_depot=True)
    entities['歸仁一課'] = build_entity(C1_ALL, '歸仁一課')
    entities['歸仁二課'] = build_entity(C2_ALL, '歸仁二課')
    for sa in SA_DISPLAY:
        course = '一課' if sa in C1_ALL else '二課'
        entities[sa] = build_entity([sa], sa)

    # 加上劉宗鑫的續保進入據點renew.ytd
    # （已在 renew sheet 裡的合計列含劉宗鑫）

    return {
        'meta': {
            'curr_month': curr_month,
            'curr_month_name': MONTH_NAMES[curr_month-1],
            'updated_at': now.strftime('%Y-%m-%d %H:%M'),
            'daily_source': daily.get('last_updated',''),
        },
        'entities': entities
    }


if __name__ == '__main__':
    # 本地測試
    print('guiren_kpi_reader.py loaded OK')
