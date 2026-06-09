import os, io, json, re
from datetime import date

SPEED_REPORT_FOLDER_ID = os.environ.get("SPEED_REPORT_FOLDER_ID", "1zUKKS0G1vsbnbptJLtU2HYTFGdIt8bv7")
DAILY_REPORT_FOLDER_ID = os.environ.get("DAILY_REPORT_FOLDER_ID", "1SDP7OJ79g6WqaoDqAQEeHZwj09MHTWyf")
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

def get_drive_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds_dict = json.loads(os.environ.get("GOOGLE_CREDENTIALS_JSON", "{}"))
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)

def get_latest_file(folder_id, keyword=""):
    service = get_drive_service()
    query = f"'{folder_id}' in parents and trashed=false"
    if keyword:
        query += f" and name contains '{keyword}'"
    results = service.files().list(
        q=query, orderBy="createdTime desc", pageSize=5,
        fields="files(id, name, createdTime)"
    ).execute()
    files = results.get("files", [])
    return files[0] if files else None

def download_file(file_id):
    from googleapiclient.http import MediaIoBaseDownload
    service = get_drive_service()
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()

def decrypt_xls(content, password):
    import msoffcrypto
    buf_in = io.BytesIO(content)
    office = msoffcrypto.OfficeFile(buf_in)
    if office.is_encrypted():
        office.load_key(password=password)
        buf_out = io.BytesIO()
        office.decrypt(buf_out)
        return buf_out.getvalue()
    return content

def find_store_cols(sh):
    targets = {"歸仁", "永康", "東台南", "西台南"}
    for r in range(sh.nrows):
        found = {}
        for c in range(1, sh.ncols - 1):
            v = str(sh.cell_value(r, c)).strip()
            if v in targets and v not in found:
                if r + 1 < sh.nrows:
                    if (str(sh.cell_value(r+1, c)).strip() == "本日" and
                        str(sh.cell_value(r+1, c+1)).strip() == "月累"):
                        found[v] = (c, c + 1)
        if "歸仁" in found and "永康" in found:
            return found, r
    return None, None

def clean_model(raw):
    """清理車款名：移除換行、多餘空白"""
    return " ".join(raw.replace("\n", " ").split()).strip()

def norm_model(raw):
    """標準化車款名用於去重"""
    return clean_model(raw).lower()

def is_valid_val(v):
    """過濾無效值：0, 0.0, "", None"""
    if v in (0, 0.0, "", None):
        return False
    return True

def get_speed_report(target_date=None):
    import xlrd
    if target_date is None:
        target_date = date.today()

    file_info = get_latest_file(SPEED_REPORT_FOLDER_ID, "業績速報")
    if not file_info:
        raise Exception(f"找不到業績速報，資料夾ID: {SPEED_REPORT_FOLDER_ID}")

    file_name = file_info["name"]
    all_dates = re.findall(r'\d{8}', file_name)
    password = all_dates[-1] if all_dates else target_date.strftime("%Y%m%d")

    content = download_file(file_info["id"])
    decrypted = decrypt_xls(content, password)
    wb = xlrd.open_workbook(file_contents=decrypted)
    sh = wb.sheet_by_index(0)

    cols, header_row = find_store_cols(sh)
    if cols is None:
        raise Exception(f"找不到據點欄位，檔案: {file_name}")

    result = {"file": file_name, "date": password}
    for name in cols:
        result[name] = {"today": {}, "mtd": {}}

    valid_labels = {"Booking", "Register", "目標達成率", "進度達成率", "來店"}
    cur_model = ""
    for r in range(header_row, sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        c1 = str(sh.cell_value(r, 1)).strip()


        # 遇到備註/保存年限/純4位日期 = 進入下一區段，停止
        if c0.startswith("備註") or c0.startswith("保存年限") or re.match(r"^\d{4}$", c0):
            break
        # 更新車款名（非空、非純數字、含括號或英文才算車款）
        if c0:
            if c0 == "TOTAL":
                cur_model = "TOTAL"
            elif not re.match(r'^[\d\.\-\/]+$', c0) and c0 not in ("上月", "本月"):
                cur_model = c0

        if c1 not in valid_labels:
            continue

        label = (cur_model + "_" + c1) if cur_model else c1

        for name, (tc, mc) in cols.items():
            tv = sh.cell_value(r, tc)
            mv = sh.cell_value(r, mc)
            if c1 in ("目標達成率", "進度達成率"):
                if isinstance(tv, float) and tv > 0: tv = f"{tv*100:.1f}%"
                if isinstance(mv, float) and mv > 0: mv = f"{mv*100:.1f}%"
            # 只保留第一次出現且有效的值
            if label not in result[name]["today"] and is_valid_val(tv):
                result[name]["today"][label] = tv
            if label not in result[name]["mtd"] and is_valid_val(mv):
                result[name]["mtd"][label] = mv

    return result


def format_speed_report_message(report):
    date_str = report.get("date", "")
    display_date = f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}" if len(date_str) == 8 else date_str

    def vt(store, key):
        try:
            v = report[store]["today"].get(key, 0)
            return int(float(v)) if v not in ("", None) else 0
        except: return 0

    def vm(store, key):
        try: return int(float(report[store]["mtd"].get(key, 0)))
        except: return 0

    def pct(store, key):
        v = report[store]["mtd"].get(key, "-")
        return v if isinstance(v, str) and "%" in v else "-"

    def sum_booking_mtd(store, display_name):
        """合計同車款（去重空白/換行變體）的月累 Booking"""
        target_norm = norm_model(display_name)
        total = 0
        for k, v in report[store]["mtd"].items():
            if not k.endswith("_Booking"): continue
            raw_name = k[:-len("_Booking")]
            if norm_model(raw_name) == target_norm:
                try: total += int(float(v))
                except: pass
        return total

    # 取所有車款（去重、過濾年份和上月等）
    seen = {}
    skip_names = {"上月", "本月", "total"}
    for store in ["歸仁", "永康", "東台南", "西台南"]:
        for k in list(report[store]["mtd"]) + list(report[store]["today"]):
            if not k.endswith("_Booking") or k == "TOTAL_Booking": continue
            raw_name = k[:-len("_Booking")]
            cn = clean_model(raw_name)
            if re.match(r'^[\d\.]+$', cn): continue  # 純數字年份
            if norm_model(raw_name) in skip_names: continue
            n = norm_model(raw_name)
            if n not in seen:
                seen[n] = cn
    all_models = sorted(seen.values())

    # 月累總覽
    gj_bk = vm("歸仁","TOTAL_Booking");  yk_bk = vm("永康","TOTAL_Booking")
    et_bk = vm("東台南","TOTAL_Booking"); wt_bk = vm("西台南","TOTAL_Booking")
    gj_rg = vm("歸仁","TOTAL_Register"); yk_rg = vm("永康","TOTAL_Register")
    et_rg = vm("東台南","TOTAL_Register"); wt_rg = vm("西台南","TOTAL_Register")
    gj_ach = pct("歸仁","TOTAL_目標達成率"); yk_ach = pct("永康","TOTAL_目標達成率")

    choc_bk = gj_bk + yk_bk; sy_bk = et_bk + wt_bk
    choc_rg = gj_rg + yk_rg; sy_rg = et_rg + wt_rg
    diff = choc_bk - sy_bk; sym = "▲" if diff >= 0 else "▼"

    # 本日
    choc_t_gj = vt("歸仁","TOTAL_Booking");  choc_t_yk = vt("永康","TOTAL_Booking")
    sy_t_et   = vt("東台南","TOTAL_Booking"); sy_t_wt  = vt("西台南","TOTAL_Booking")

    def sum_today(store, display_name):
        target_norm = norm_model(display_name)
        total = 0
        for k, v in report[store]["today"].items():
            if not k.endswith("_Booking"): continue
            if norm_model(k[:-len("_Booking")]) == target_norm:
                try: total += int(float(v))
                except: pass
        return total

    # 本日車型（巧克力）
    choc_today_lines = []
    for mn in all_models:
        gj_t = sum_today("歸仁", mn); yk_t = sum_today("永康", mn)
        if gj_t + yk_t > 0:
            choc_today_lines.append(f"  {mn}：歸仁{gj_t} / 永康{yk_t}")

    # 本日車型（伸陽）
    sy_today_lines = []
    for mn in all_models:
        et_t = sum_today("東台南", mn); wt_t = sum_today("西台南", mn)
        if et_t + wt_t > 0:
            sy_today_lines.append(f"  {mn}：東台南{et_t} / 西台南{wt_t}")

    choc_today_sec = "\n".join(choc_today_lines) if choc_today_lines else "  （無訂單）"
    sy_today_sec   = "\n".join(sy_today_lines)   if sy_today_lines   else "  （無訂單）"

    # 月累車型明細
    mtd_lines = []
    for mn in all_models:
        gj_m = sum_booking_mtd("歸仁",  mn); yk_m = sum_booking_mtd("永康",  mn)
        et_m = sum_booking_mtd("東台南", mn); wt_m = sum_booking_mtd("西台南", mn)
        if gj_m + yk_m + et_m + wt_m > 0:
            mtd_lines.append(
                f"  {mn}\n"
                f"    🍫 歸{gj_m}/永{yk_m}={gj_m+yk_m}台　"
                f"⚔️ 東{et_m}/西{wt_m}={et_m+wt_m}台"
            )
    mtd_sec = "\n".join(mtd_lines) if mtd_lines else "  （無資料）"

    lines = [
        f"📊 {display_date} 業績速報",
        "",
        "━━━━━━━━━━━━━━",
        "🍫 巧克力 月累",
        f"  歸仁{gj_bk}台 永康{yk_bk}台 合計{choc_bk}台",
        f"  達成率 歸仁{gj_ach} 永康{yk_ach}",
        "",
        "⚔️ 伸陽 月累",
        f"  東台南{et_bk}台 西台南{wt_bk}台 合計{sy_bk}台",
        "",
        f"📌 差距：{sym}{abs(diff)}台",
        "━━━━━━━━━━━━━━",
        "🚗 本日訂單",
        f"🍫 巧克力 {choc_t_gj+choc_t_yk}台（歸仁{choc_t_gj} / 永康{choc_t_yk}）",
        choc_today_sec,
        "",
        f"⚔️ 伸陽 {sy_t_et+sy_t_wt}台（東台南{sy_t_et} / 西台南{sy_t_wt}）",
        sy_today_sec,
        "━━━━━━━━━━━━━━",
        "📊 月累車型明細",
        mtd_sec,
        "━━━━━━━━━━━━━━",
        f"🏁 領牌月累",
        f"  🍫 {choc_rg}台（歸{gj_rg}/永{yk_rg}）　⚔️ {sy_rg}台（東{et_rg}/西{wt_rg}）",
    ]
    return "\n".join(lines)


def get_daily_report(target_date=None):
    import xlrd
    if target_date is None:
        target_date = date.today()

    file_info = get_latest_file(DAILY_REPORT_FOLDER_ID, "歸仁日報表")
    if not file_info:
        file_info = get_latest_file(DAILY_REPORT_FOLDER_ID, "日報表")
    if not file_info:
        raise Exception(f"找不到日報表，資料夾ID: {DAILY_REPORT_FOLDER_ID}")

    file_name = file_info["name"]
    content = download_file(file_info["id"])
    members = ["林定緯", "林適緯", "陳建道", "陳星佑", "張姉瑀", "歐陽文智", "蔡明憬"]
    result = {"file": file_name, "data": {}}

    if file_name.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = None
        for name in wb.sheetnames:
            if str(date.today().month) in name:
                ws = wb[name]; break
        if not ws: ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                name = str(cell).strip() if cell else ""
                if name in members:
                    result["data"][name] = str([v for v in row if v is not None][:15]); break
        wb.close()
    else:
        wb = xlrd.open_workbook(file_contents=content)
        sh = wb.sheet_by_index(0)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = str(sh.cell_value(r, c)).strip()
                if v in members and v not in result["data"]:
                    result["data"][v] = str([sh.cell_value(r, cc) for cc in range(min(sh.ncols, 15))])
    return result
