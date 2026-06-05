import os
import io
import json
import requests
from datetime import datetime, date, timedelta, timezone

LINE_TOKEN        = os.environ.get("LINE_TOKEN")
TARGET_GROUP_ID   = os.environ.get("REMINDER_GROUP_ID")
RENEWAL_FILE_ID   = os.environ.get("RENEWAL_SHEET_ID", "1-6Wmly1lKSLVOEUspwcV9TPsghdjyMC_")
REMIND_START_HOUR = 9
REMIND_END_HOUR   = 20
STATE_FILE        = "reminder_state.json"
MEMBER_USER_IDS   = json.loads(os.environ.get("MEMBER_USER_IDS", "{}"))
TW_TZ             = timezone(timedelta(hours=8))

_holiday_cache = {}

def get_holidays(year: int) -> set:
    if year in _holiday_cache:
        return _holiday_cache[year]
    holidays = set()
    try:
        url = f"https://data.ntpc.gov.tw/api/datasets/308DCD75-6119-4125-8843-2057C0E43ED5/json?$top=500&$filter=year%20eq%20{year}"
        resp = requests.get(url, timeout=10, verify=False)
        if resp.status_code == 200:
            try:
                data = resp.json()
                for item in data:
                    d = item.get("date", "")
                    if len(d) == 8:
                        holidays.add(date(int(d[:4]), int(d[4:6]), int(d[6:])))
            except:
                pass
    except Exception as e:
        print(f"[WARNING] 假日 API 失敗: {e}")
    _holiday_cache[year] = holidays
    return holidays

def is_off_day(d: date) -> bool:
    if d.weekday() == 6:
        return True
    return d in get_holidays(d.year)

def next_work_day(d: date) -> date:
    next_d = d
    while is_off_day(next_d):
        next_d += timedelta(days=1)
    return next_d

def get_trigger_date(month_day: int, year: int, month: int) -> date:
    try:
        original = date(year, month, month_day)
    except ValueError:
        return None
    return next_work_day(original)

def get_days_left_in_month(d: date) -> int:
    if d.month == 12:
        last_day = date(d.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(d.year, d.month + 1, 1) - timedelta(days=1)
    return (last_day - d).days + 1

# ── 讀取續保資料（Drive API 下載 xlsx）──────────────
def get_renewal_data() -> dict:
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
        import openpyxl

        creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
        if not creds_json:
            print("[ERROR] 缺少 GOOGLE_CREDENTIALS_JSON")
            return {}

        creds_dict = json.loads(creds_json)
        scopes = ["https://www.googleapis.com/auth/drive.readonly"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        service = build("drive", "v3", credentials=creds)

        request = service.files().get_media(fileId=RENEWAL_FILE_ID)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)

        today = date.today()
        ws = None
        for name in wb.sheetnames:
            if f"{today.month:02d}月" in name or f"{today.month}月" in name:
                ws = wb[name]
                break
        if not ws:
            ws = wb.active

        skip_names = {"劉珈微", "劉宗鑫", "歸仁一課", "歸仁二課",
                      "合計", "營業員", "None", ""}

        def safe_int(val):
            try:
                if val is None or str(val).strip() in ("", "-", "None"):
                    return 0
                return int(float(str(val).replace(",", "").strip()))
            except:
                return 0

        result = {}
        headers = []
        header_found = False

        for row in ws.iter_rows(values_only=True):
            row_vals = [str(v).strip() if v is not None else "" for v in row]
            if not header_found:
                if "營業員" in row_vals and "母數" in row_vals:
                    headers = row_vals
                    header_found = True
                continue

            name = row_vals[0] if row_vals else ""
            if not name or name in skip_names:
                continue

            def col(keyword):
                for i, h in enumerate(headers):
                    if keyword in h:
                        return safe_int(row_vals[i]) if i < len(row_vals) else 0
                return 0

            mu = col("母數")
            if mu == 0:
                continue

            result[name] = {
                "母數":     mu,
                "預估":     col("預估"),
                "已收":     col("已收"),
                "續保率":   col("已收") / mu if mu > 0 else 0,
                "首年母數": col("首年續保母") or col("首年母"),
                "首年預估": col("首年續保預") or col("首年預"),
                "首年已收": col("首年續保已") or col("首年已"),
                "車體母數": col("車體母"),
                "車體預估": col("車體預"),
                "車體已收": col("車體已"),
            }

        wb.close()
        print(f"[OK] 讀取續保資料成功，共 {len(result)} 人")
        return result

    except Exception as e:
        print(f"[ERROR] 讀取續保資料失敗: {e}")
        return {}

# ── 提醒條件判斷 ────────────────────────────────────
def should_remind(person_data: dict, today: date):
    day    = today.day
    mu     = person_data["母數"]
    est    = person_data["預估"]
    col    = person_data["已收"]
    rate   = person_data["續保率"]

    if mu == 0:
        return False, ""

    y, m = today.year, today.month
    t1 = get_trigger_date(1,  y, m)
    t2 = get_trigger_date(7,  y, m)
    t3 = get_trigger_date(15, y, m)
    t4 = get_trigger_date(20, y, m)

    if today == t1 and rate < 0.20:
        gap = max(0, round(mu * 0.20) - col)
        return True, f"月初進度僅 {rate*100:.1f}%（{col}/{mu}台），距門檻20%還差 {gap} 台"
    if today == t2 and rate < 0.50:
        gap = max(0, round(mu * 0.50) - col)
        return True, f"進度僅 {rate*100:.1f}%（{col}/{mu}台），距門檻50%還差 {gap} 台"
    if today == t3 and est > 0 and col < est:
        return True, f"進度 {rate*100:.1f}%（{col}/{mu}台），落後預估 {est-col} 台"
    if today == t4 and est > 0 and col < est:
        return True, f"⚠️ 第二次警示：進度 {rate*100:.1f}%（{col}/{mu}台），落後預估 {est-col} 台"
    if day >= 21 and est > 0 and col < est:
        days_left = get_days_left_in_month(today)
        return True, f"距月底剩 {days_left} 天，進度 {rate*100:.1f}%（{col}/{mu}台），還差 {est-col} 台"

    return False, ""

# ── LINE 發送 ────────────────────────────────────────
def push_mention(name: str, reason: str, is_followup: bool = False):
    user_id = MEMBER_USER_IDS.get(name)
    month   = date.today().month

    if is_followup:
        text = f"@{name} 尚未收到你的回覆\n請盡快說明 {month}月續保進度狀況 🙏"
    else:
        text = f"📊 {month}月續保進度提醒\n@{name}\n{reason}\n\n請今天內回覆追蹤計畫 🙏"

    msg = {"type": "text", "text": text}

    if user_id:
        at = f"@{name}"
        idx = text.index(at)
        msg["mention"] = {
            "mentionees": [{
                "index": idx,
                "length": len(at),
                "type": "user",
                "userId": user_id
            }]
        }

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Authorization": f"Bearer {LINE_TOKEN}",
        "Content-Type": "application/json"
    }
    resp = requests.post(url, headers=headers,
                         json={"to": TARGET_GROUP_ID, "messages": [msg]})
    print(f"[LINE] @{name} 發送: {resp.status_code}")

# ── 狀態管理 ─────────────────────────────────────────
def load_state() -> dict:
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def save_state(state: dict):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def reset_daily_state(state: dict, today_str: str) -> dict:
    for name in state:
        if state[name].get("last_reset") != today_str:
            state[name]["replied_today"]  = False
            state[name]["reminded_today"] = False
            state[name]["last_reset"]     = today_str
    return state

# ── 主入口 ───────────────────────────────────────────
def run_reminder():
    now          = datetime.now(TW_TZ)
    today        = now.date()
    today_str    = today.isoformat()
    current_hour = now.hour

    if current_hour < REMIND_START_HOUR or current_hour >= REMIND_END_HOUR:
        print(f"[SKIP] 目前台灣時間 {current_hour}:00，不在提醒時間內（{REMIND_START_HOUR}-{REMIND_END_HOUR}）")
        return

    if today.weekday() == 6:
        print("[SKIP] 今天是週日，跳過")
        return

    if today in get_holidays(today.year):
        print("[SKIP] 今天是國定假日，跳過")
        return

    renewal_data = get_renewal_data()
    if not renewal_data:
        print("[SKIP] 無法取得續保資料")
        return

    state = load_state()
    state = reset_daily_state(state, today_str)

    for name, data in renewal_data.items():
        if name not in state:
            state[name] = {
                "replied_today":  False,
                "reminded_today": False,
                "last_reset":     today_str,
                "last_remind_time": None
            }

        person = state[name]
        if person.get("replied_today"):
            continue

        need_remind, reason = should_remind(data, today)
        if not need_remind:
            continue

        last_remind = person.get("last_remind_time")

        if not person.get("reminded_today"):
            if current_hour == REMIND_START_HOUR:
                push_mention(name, reason, is_followup=False)
                person["reminded_today"]    = True
                person["last_remind_time"]  = now.isoformat()
                print(f"[REMIND] 首次提醒 @{name}")
        else:
            if last_remind:
                last_dt     = datetime.fromisoformat(last_remind)
                hours_since = (now - last_dt).total_seconds() / 3600
                if hours_since >= 2:
                    push_mention(name, reason, is_followup=True)
                    person["last_remind_time"] = now.isoformat()
                    print(f"[FOLLOWUP] 追蹤提醒 @{name}（距上次 {hours_since:.1f} 小時）")

    save_state(state)
    print(f"[DONE] {now.strftime('%Y-%m-%d %H:%M')} 台灣時間，提醒任務完成")

def mark_replied(user_display_name: str):
    state     = load_state()
    today_str = date.today().isoformat()
    for name in state:
        if name in user_display_name or user_display_name in name:
            state[name]["replied_today"] = True
            save_state(state)
            print(f"[REPLIED] {name} 已回覆")
            return True
    return False
