"""
歸仁儀表板 - 資料解析模組（v4）
v4 改動：YoY 去年同期改用「年度各月累計 + 當月 sheet 月累」組合，
解決原本 114年度累計報表不逐日更新導致去年同期偏低的問題。
"""

import io
import os
import re
import json
from datetime import date, datetime
from collections import defaultdict

import openpyxl

from drive_reader import download_file, DAILY_REPORT_FOLDER_ID

DATA_FILE = "dashboard_data.json"
MONTHS = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']


def _parse_filename_date(filename):
    m = re.search(r'\d{3}\s*([0-9\-\s]+)', filename)
    if not m:
        return (0, 0)
    raw = m.group(1)
    raw = raw.replace(' ', '').strip('-')
    if not raw:
        return (0, 0)
    parts = raw.split('-')
    try:
        if len(parts) == 1:
            digits = parts[0]
            if len(digits) < 3:
                return (0, 0)
            if len(digits) == 3:
                mm, dd = digits[0], digits[1:]
            else:
                mm, dd = digits[:2], digits[2:4]
            return (int(mm), int(dd))
        else:
            end = parts[-1]
            start = parts[0]
            if len(end) >= 4:
                mm, dd = end[:2], end[2:4]
            elif len(end) <= 2:
                mm = start[:2] if len(start) >= 3 else start[:1]
                dd = end
            else:
                mm, dd = end[:1], end[1:]
            return (int(mm), int(dd))
    except (ValueError, IndexError):
        return (0, 0)


def find_latest_115_file(folder_id):
    from drive_reader import get_drive_service
    service = get_drive_service()
    query = f"'{folder_id}' in parents and name contains '歸仁日報表115' and trashed = false"
    resp = service.files().list(q=query, pageSize=200, fields="files(id, name)").execute()
    files = resp.get("files", [])
    if not files:
        return None
    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date.sort(key=lambda x: x[1], reverse=True)
    return files_with_date[0][0]


def find_closest_114_file(folder_id, target_month, target_day):
    from drive_reader import get_drive_service
    service = get_drive_service()
    query = f"'{folder_id}' in parents and name contains '歸仁日報表114' and trashed = false"
    resp = service.files().list(q=query, pageSize=300, fields="files(id, name)").execute()
    files = resp.get("files", [])
    if not files:
        return None, "Drive查詢114年檔案回傳空清單"

    def day_ord(mm, dd):
        return mm * 31 + dd

    target_ord = day_ord(target_month, target_day)
    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date = [(f, md) for f, md in files_with_date if md[0] != 0]
    if not files_with_date:
        return None, "114年檔案都解析不出日期: " + ",".join(f["name"] for f in files[:5])
    files_with_date.sort(key=lambda x: abs(day_ord(*x[1]) - target_ord))
    best_file, (mm, dd) = files_with_date[0]
    return (best_file, mm, dd), None


def sheet_to_grid(wb, sheet_name, max_cols=210):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    grid = []
    for row in ws.iter_rows(min_col=1, max_col=max_cols, values_only=True):
        grid.append(row)
    return grid


def grid_get(grid, row, col):
    r_idx = row - 1
    c_idx = col - 1
    if r_idx < 0 or r_idx >= len(grid):
        return None
    row_tuple = grid[r_idx]
    if c_idx < 0 or c_idx >= len(row_tuple):
        return None
    return row_tuple[c_idx]


def is_person_name(name):
    BLACKLIST_SUBSTR = ['合計', '月累', '課', '歸一', '歸二', '歸三', '歸仁', '公司']
    if not name or not str(name).strip():
        return False
    name = str(name).strip()
    if any(s in name for s in BLACKLIST_SUBSTR):
        return False
    if len(name) < 2 or len(name) > 5:
        return False
    return True


def normalize_model(name):
    if name is None:
        return None
    n = str(name).strip().upper().replace(' ', '').replace('\n', '')
    mapping = {
        'ZRV': 'ZRV', 'CR-V': 'CR-V', 'CRV': 'CR-V', 'HRV': 'HRV', 'HR-V': 'HRV',
        'FIT': 'FIT', 'CIVIC': 'CIVIC', 'ODYSSEY': 'ODYSSEY', 'PRELUDE': 'PRELUDE',
        'ACCORD': 'ACCORD', 'INSIGHT': 'INSIGHT', 'CR-Z': 'CR-Z', 'CRZ': 'CR-Z',
    }
    return mapping.get(n, n)


def find_groups(grid, header_row, sub_row, start_col, end_col_exclusive):
    starts = []
    for c in range(start_col, end_col_exclusive):
        v = grid_get(grid, header_row, c)
        if v is not None and str(v).strip():
            starts.append((c, str(v).strip()))
    groups = []
    for i, (c, name) in enumerate(starts):
        next_c = starts[i + 1][0] if i + 1 < len(starts) else end_col_exclusive
        cumcol = None
        for cc in range(c, next_c):
            sv = grid_get(grid, sub_row, cc)
            if sv is not None and '月累' in str(sv):
                cumcol = cc
                break
        if cumcol is not None:
            groups.append((normalize_model(name), c, cumcol))
    return groups


def parse_month_sheet(wb, sheet_name):
    grid = sheet_to_grid(wb, sheet_name, max_cols=80)
    if not grid:
        return {}
    reg_groups = find_groups(grid, 4, 5, 2, 36)
    ord_groups = find_groups(grid, 4, 5, 49, 76)
    result = {}
    for r in range(6, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        result.setdefault(name, {})
        for model, gc, cumcol in reg_groups:
            v = grid_get(grid, r, cumcol)
            v = v if isinstance(v, (int, float)) else 0
            result[name].setdefault(model, {}).setdefault('領牌', 0)
            result[name][model]['領牌'] += v
        for model, gc, cumcol in ord_groups:
            v = grid_get(grid, r, cumcol)
            v = v if isinstance(v, (int, float)) else 0
            result[name].setdefault(model, {}).setdefault('訂單', 0)
            result[name][model]['訂單'] += v
    return result


CRV_TRIM_ORDER_COLS = {'CR-V(PET)': [49, 50], 'CR-V(e:HEV)': [51, 52]}


def parse_crv_trim_split(wb, sheet_name):
    grid = sheet_to_grid(wb, sheet_name, max_cols=80)
    if not grid:
        return {}
    sub_row = 5
    label_51 = grid_get(grid, sub_row, 51)
    label_52 = grid_get(grid, sub_row, 52)
    is_new_layout = (
        label_51 and 'HEV' in str(label_51).upper()
        and label_52 and 'HEV' in str(label_52).upper()
    )
    if not is_new_layout:
        return {}
    result = {}
    for r in range(6, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        result.setdefault(name, {})
        for label, cols in CRV_TRIM_ORDER_COLS.items():
            total = 0
            for c in cols:
                v = grid_get(grid, r, c)
                if isinstance(v, (int, float)):
                    total += v
            result[name][label] = total
    return result


def parse_year_summary(wb, sheet_name):
    """讀年度累計報表，回傳 {person: 年度累計領牌}（col27）。保留供相容。"""
    grid = sheet_to_grid(wb, sheet_name, max_cols=30)
    if not grid:
        return {}
    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        val = grid_get(grid, r, 27)
        if isinstance(val, (int, float)):
            result[name] = val
    return result


def parse_year_by_month(wb, sheet_name, up_to_month):
    """讀114年度sheet各月領牌加總（1 ~ up_to_month）。
    第N月領牌欄 = col(1 + N*2)，確認：1月=col3, 2月=col5。
    回傳 {person: 加總領牌}
    """
    grid = sheet_to_grid(wb, sheet_name, max_cols=30)
    if not grid:
        return {}
    result = {}
    for r in range(1, len(grid) + 1):
        name = grid_get(grid, r, 1)
        if not is_person_name(name):
            continue
        name = str(name).strip()
        total = 0
        for m in range(1, up_to_month + 1):
            col = 1 + m * 2
            v = grid_get(grid, r, col)
            total += int(v) if isinstance(v, (int, float)) else 0
        result[name] = total
    return result


def parse_month_reg_total(wb, sheet_name):
    """讀某月 sheet 的月累計領牌，回傳 {person: 領牌月累}。
    用於 YoY：當年度累計報表當月數字為 0 時，補上當月 sheet 的即時數字。
    """
    r = parse_month_sheet(wb, sheet_name)
    result = {}
    for person, models in r.items():
        total = sum(v.get('領牌', 0) for v in models.values())
        if total > 0:
            result[person] = total
    return result


HISTORY_DRIVE_FILENAME = "dashboard_order_history.json"
SEED_HISTORY_FILE = "order_history_seed.json"


def load_order_history():
    from drive_json_store import load_json_from_drive
    history = {}
    if os.path.exists(SEED_HISTORY_FILE):
        with open(SEED_HISTORY_FILE, "r", encoding="utf-8") as f:
            history.update(json.load(f))
    history.update(load_json_from_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME))
    for snap in history.values():
        for models in snap.values():
            models.pop('CR-V', None)
    return history


def save_order_history(history):
    from drive_json_store import save_json_to_drive
    save_json_to_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME, history)


DAILY_VALUE_MODELS = {'CR-V(PET)', 'CR-V(e:HEV)'}


def compute_last_order_tracking(history, today_str):
    dates_sorted = sorted(history.keys())
    people = set()
    models = set()
    for snap in history.values():
        for p, md in snap.items():
            people.add(p)
            models.update(md.keys())

    from collections import OrderedDict
    by_month = OrderedDict()
    for d in dates_sorted:
        by_month.setdefault(d[:7], []).append(d)

    today = date.fromisoformat(today_str)
    result = {}
    for p in people:
        result[p] = {}
        for model in models:
            last_date = None
            if model in DAILY_VALUE_MODELS:
                for d in dates_sorted:
                    cur_val = history[d].get(p, {}).get(model, 0)
                    if cur_val > 0:
                        last_date = d
            else:
                for ym, ds in by_month.items():
                    prev_val = None
                    for d in ds:
                        cur_val = history[d].get(p, {}).get(model, 0)
                        if prev_val is None:
                            if cur_val > 0:
                                last_date = d
                        else:
                            if cur_val > prev_val:
                                last_date = d
                        prev_val = cur_val
            if last_date is None:
                continue
            days_since = (today - date.fromisoformat(last_date)).days
            result[p][model] = {"last_order_date": last_date, "days_since": days_since, "approx": False}
    return result


MONTH_LAST_DAY_2026 = {
    '1月': (1, 31), '2月': (2, 28), '3月': (3, 31), '4月': (4, 30),
    '5月': (5, 31), '6月': (6, 30), '7月': (7, 31), '8月': (8, 31),
    '9月': (9, 30), '10月': (10, 31), '11月': (11, 30), '12月': (12, 31),
}


def fill_fallback_from_monthly(last_order_tracking, month_sheets_cache, months_to_sum, today_str):
    today = date.fromisoformat(today_str)
    people_models_with_data = set()
    for m in months_to_sum:
        for p, models in month_sheets_cache.get(m, {}).items():
            for model, vals in models.items():
                if vals.get('訂單', 0) > 0:
                    people_models_with_data.add((p, model))

    for p, model in people_models_with_data:
        if model in DAILY_VALUE_MODELS:
            continue
        existing = last_order_tracking.get(p, {}).get(model)
        if existing is not None:
            continue
        for m in reversed(months_to_sum):
            v = month_sheets_cache.get(m, {}).get(p, {}).get(model, {}).get('訂單', 0)
            if v > 0:
                mm, dd = MONTH_LAST_DAY_2026[m]
                candidate = date(today.year, mm, dd)
                if candidate > today:
                    candidate = today
                last_order_tracking.setdefault(p, {})
                last_order_tracking[p][model] = {
                    "last_order_date": candidate.isoformat(),
                    "days_since": (today - candidate).days,
                    "approx": True,
                }
                break
    return last_order_tracking


TEAM_ORDER = ['林定緯', '林適緯', '陳建道', '陳星佑', '張姉瑀', '歐陽文智', '蔡明憬']

TEAM_STRUCTURE = {
    '歸仁一課': ['林定緯', '林適緯', '陳建道'],
    '歸仁二課': ['陳星佑', '張姉瑀', '歐陽文智', '蔡明憬'],
}
PERSON_TO_DEPT = {p: dept for dept, members in TEAM_STRUCTURE.items() for p in members}
PERSON_TO_DEPT['劉珈微'] = '歸仁一課'


def compute_dept_totals_scalar(person_value_dict):
    totals = {dept: 0 for dept in TEAM_STRUCTURE}
    for p, v in person_value_dict.items():
        dept = PERSON_TO_DEPT.get(p)
        if dept:
            totals[dept] += v
    return totals


def compute_dept_totals_by_model(person_model_dict):
    totals = {dept: defaultdict(int) for dept in TEAM_STRUCTURE}
    for p, models in person_model_dict.items():
        dept = PERSON_TO_DEPT.get(p)
        if dept:
            for model, v in models.items():
                totals[dept][model] += v
    return {dept: dict(models) for dept, models in totals.items()}


def sort_by_team_order(d):
    def key(item):
        name = item[0]
        try:
            return (0, TEAM_ORDER.index(name))
        except ValueError:
            return (1, name)
    return dict(sorted(d.items(), key=key))


def sort_by_total_desc(d):
    return dict(sorted(d.items(), key=lambda item: sum(item[1].values()) if isinstance(item[1], dict) else item[1], reverse=True))


def build_daily_snapshot(wb, month_key):
    r = parse_month_sheet(wb, month_key)
    crv_split = parse_crv_trim_split(wb, month_key)
    snapshot = {}
    for person, models in r.items():
        snapshot[person] = {k: v.get('訂單', 0) for k, v in models.items()
                             if '訂單' in v and k != 'CR-V'}
    for person, split_vals in crv_split.items():
        snapshot.setdefault(person, {})
        for label, val in split_vals.items():
            snapshot[person][label] = val
    return snapshot


def reset_order_history():
    from drive_json_store import save_json_to_drive
    save_json_to_drive(DAILY_REPORT_FOLDER_ID, HISTORY_DRIVE_FILENAME, {})
    return {"status": "ok", "message": "Drive上的歷史紀錄已清空，種子檔不受影響"}


def backfill_full_history(max_seconds=100, max_files=30):
    import time
    import gc
    start_time = time.time()

    from drive_reader import get_drive_service, download_file
    service = get_drive_service()
    query = f"'{DAILY_REPORT_FOLDER_ID}' in parents and name contains '歸仁日報表115' and trashed = false"
    resp = service.files().list(q=query, pageSize=300, fields="files(id, name)").execute()
    files = resp.get("files", [])

    files_with_date = [(f, _parse_filename_date(f["name"])) for f in files]
    files_with_date = [(f, md) for f, md in files_with_date if md[0] != 0]
    files_with_date.sort(key=lambda x: x[1])

    history = load_order_history()
    processed = 0
    errors = []
    timed_out = False

    for f, (mm, dd) in files_with_date:
        date_str = f"2026-{mm:02d}-{dd:02d}"
        if date_str in history:
            continue
        if processed >= max_files or time.time() - start_time > max_seconds:
            timed_out = True
            break
        try:
            content = download_file(f["id"])
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            month_key = f"{mm}月"
            if month_key in wb.sheetnames:
                history[date_str] = build_daily_snapshot(wb, month_key)
                processed += 1
            wb.close()
            del content, wb
            gc.collect()
        except Exception as e:
            errors.append(f"{f['name']}: {str(e)[:100]}")

    save_order_history(history)

    return {
        "processed_this_run": processed,
        "total_files_found": len(files_with_date),
        "dates_now_in_history": len(history),
        "timed_out": timed_out,
        "done": not timed_out,
        "errors": errors[:10],
    }


def build_dashboard_data():
    file_info = find_latest_115_file(DAILY_REPORT_FOLDER_ID)
    if not file_info:
        raise RuntimeError("Drive 資料夾裡找不到115年歸仁日報表檔案")

    content = download_file(file_info["id"])
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    today = date.today()
    current_month_key = f"{today.month}月"

    ytd = defaultdict(lambda: defaultdict(lambda: {'領牌': 0, '訂單': 0}))
    months_to_sum = [m for m in MONTHS if m in wb.sheetnames]
    month_sheets_cache = {}
    month_sheets_cache_for_tracking = {}
    for m in months_to_sum:
        r = parse_month_sheet(wb, m)
        month_sheets_cache[m] = r
        for person, models in r.items():
            for model, vals in models.items():
                ytd[person][model]['領牌'] += vals.get('領牌', 0)
                ytd[person][model]['訂單'] += vals.get('訂單', 0)

        crv_split = parse_crv_trim_split(wb, m)
        r_tracking = {}
        for person, models in r.items():
            r_tracking[person] = {k: v for k, v in models.items() if k != 'CR-V'}
        for person, split_vals in crv_split.items():
            r_tracking.setdefault(person, {})
            for label, val in split_vals.items():
                r_tracking[person][label] = {'訂單': val}
        month_sheets_cache_for_tracking[m] = r_tracking

    item1 = {p: sum(v['領牌'] for v in models.values()) for p, models in ytd.items()}
    team_total_ytd_registration = sum(item1.values())
    item1_dept_totals = compute_dept_totals_scalar(item1)

    item4 = {p: {mo: v['領牌'] for mo, v in models.items() if v['領牌'] > 0}
             for p, models in ytd.items()}
    item4 = sort_by_total_desc(item4)

    cur = month_sheets_cache.get(current_month_key, {})
    cur_tracking = month_sheets_cache_for_tracking.get(current_month_key, {})
    month_progress = {}
    for p, models in cur.items():
        month_progress[p] = {
            '訂單': sum(v.get('訂單', 0) for v in models.values()),
            '領牌': sum(v.get('領牌', 0) for v in models.values()),
        }

    today_str = today.isoformat()
    history = load_order_history()
    if cur_tracking:
        snapshot = {p: {m: v.get('訂單', 0) for m, v in models.items() if '訂單' in v}
                    for p, models in cur_tracking.items()}
        history[today_str] = snapshot
        save_order_history(history)
    last_order_tracking = compute_last_order_tracking(history, today_str)
    last_order_tracking = fill_fallback_from_monthly(
        last_order_tracking, month_sheets_cache_for_tracking, months_to_sum, today_str
    )
    last_order_tracking = sort_by_team_order(last_order_tracking)

    wb.close()
    del content

    # ── YoY：去年同期 = 年度各月累計（上月底）+ 當月 sheet 月累（即時）──
    yoy_comparison = None
    try:
        found, err = find_closest_114_file(DAILY_REPORT_FOLDER_ID, today.month, today.day)
        if found is None:
            yoy_comparison = {"error": err}
        else:
            last_year_file, ly_month, ly_day = found
            ly_content = download_file(last_year_file["id"])
            ly_wb = openpyxl.load_workbook(io.BytesIO(ly_content), data_only=True, read_only=True)

            # 1. 年度sheet讀上月底（1 ~ curr_month-1）各月領牌加總
            prev_month = today.month - 1
            if prev_month >= 1:
                ly_prev = parse_year_by_month(ly_wb, "114年度", prev_month)
            else:
                ly_prev = {}

            # 2. 年度sheet讀當月數字
            curr_col = 1 + today.month * 2
            ly_ytd_grid = sheet_to_grid(ly_wb, "114年度", max_cols=30)
            curr_from_ytd = {}
            for r in range(1, len(ly_ytd_grid) + 1):
                name = grid_get(ly_ytd_grid, r, 1)
                if not is_person_name(name):
                    continue
                name = str(name).strip()
                v = grid_get(ly_ytd_grid, r, curr_col)
                curr_from_ytd[name] = int(v) if isinstance(v, (int, float)) else 0

            # 3. 當月 sheet 月累（備援）
            ly_curr_sheet = f"{today.month}月"
            ly_curr_reg = {}
            if ly_curr_sheet in ly_wb.sheetnames:
                ly_curr_reg = parse_month_reg_total(ly_wb, ly_curr_sheet)

            ly_wb.close()
            del ly_content

            # 4. 合併：上月累計 + 當月（年度sheet有值優先，否則用當月sheet）
            all_people = set(ly_prev.keys()) | set(curr_from_ytd.keys()) | set(ly_curr_reg.keys())
            ly_combined = {}
            for p in all_people:
                prev = ly_prev.get(p, 0)
                curr_ytd = curr_from_ytd.get(p, 0)
                curr_sheet = ly_curr_reg.get(p, 0)
                curr_final = curr_ytd if curr_ytd > 0 else curr_sheet
                ly_combined[p] = prev + curr_final

            # 手動校正：陳星佑114年1~3月在永康，歸仁報表看不到
            KNOWN_CORRECTIONS = {"陳星佑": 12}
            for _name, _add in KNOWN_CORRECTIONS.items():
                if _name in ly_combined:
                    ly_combined[_name] += _add

            yoy_comparison = {
                "last_year_file": last_year_file["name"],
                "last_year_date": f"2025-{ly_month:02d}-{ly_day:02d}",
                "last_year_ytd": {k: int(v) for k, v in ly_combined.items()},
                "yoy_method": f"年度各月累計(1~{prev_month}月) + 當月sheet即時領牌",
            }
    except Exception as e:
        import traceback
        yoy_comparison = {"error": str(e), "trace": traceback.format_exc()[-500:]}

    EXCLUDE_FROM_PERSONAL = {'劉珈微'}
    item1 = {p: v for p, v in item1.items() if p not in EXCLUDE_FROM_PERSONAL}
    item4 = {p: v for p, v in item4.items() if p not in EXCLUDE_FROM_PERSONAL}
    month_progress = {p: v for p, v in month_progress.items() if p not in EXCLUDE_FROM_PERSONAL}
    last_order_tracking = {p: v for p, v in last_order_tracking.items() if p not in EXCLUDE_FROM_PERSONAL}

    item4_dept_totals = compute_dept_totals_by_model(item4)
    month_progress_dept_totals = compute_dept_totals_by_model(month_progress)
    yoy_last_year_dept_totals = None
    yoy_this_year_dept_totals = None
    if yoy_comparison and yoy_comparison.get("last_year_ytd") is not None:
        yoy_last_year_dept_totals = compute_dept_totals_scalar(yoy_comparison["last_year_ytd"])
        yoy_this_year_dept_totals = item1_dept_totals

    data = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": file_info["name"],
        "team_structure": TEAM_STRUCTURE,
        "item1_ytd_registration": item1,
        "item1_dept_totals": item1_dept_totals,
        "team_total_ytd_registration": team_total_ytd_registration,
        "item4_ytd_by_model": item4,
        "item4_dept_totals": item4_dept_totals,
        "month_progress": month_progress,
        "month_progress_dept_totals": month_progress_dept_totals,
        "last_order_tracking": last_order_tracking,
        "yoy_comparison": yoy_comparison,
        "yoy_last_year_dept_totals": yoy_last_year_dept_totals,
        "yoy_this_year_dept_totals": yoy_this_year_dept_totals,
    }

    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return data


if __name__ == "__main__":
    result = build_dashboard_data()
    print(json.dumps(result, ensure_ascii=False, indent=2))
